import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
import pyautogui
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import logging
import threading
import keyboard  # Ensure you have the `keyboard` package installed (pip install keyboard)
import cv2  # Ensure you have the `opencv-python` package installed (pip install opencv-python)
import win32api
import win32gui

# Configure logging
logging.basicConfig(filename='automation_log.txt', level=logging.DEBUG, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Global variable to store data
data = None
file_path = "automate.xlsx"

# Function to load data from an Excel file
def load_data():
    global data
    if os.path.exists(file_path):
        try:
            data = pd.read_excel(file_path, header=None)  # Read without headers
            logging.debug(f"Loaded data:\n{data}")  # Debug: Log the entire loaded data

            if data.empty:
                data = pd.DataFrame(columns=['Date', 'Invoice Number', 'Serial'])
            else:
                if len(data.columns) == 2:
                    data.columns = ['Date', 'Invoice Number']
                    data['Serial'] = ''
                elif len(data.columns) == 3:
                    data.columns = ['Date', 'Invoice Number', 'Serial']
                else:
                    raise ValueError("Unexpected number of columns in the loaded data.")

            logging.debug(f"Columns: {data.columns}")  # Debug: Log the column names
            update_listbox(data)
        except Exception as e:
            logging.error(f"Error loading data: {e}")  # Debug
            messagebox.showerror("Error", f"Failed to load data: {e}")
    else:
        data = pd.DataFrame(columns=['Date', 'Invoice Number', 'Serial'])

# Function to check for the pop-up and close it if it appears
def check_and_close_popup(image_path, retries=5, interval=1):
    for _ in range(retries):
        try:
            popup_location = pyautogui.locateOnScreen(image_path, confidence=0.9)
            if popup_location:
                logging.debug("Pop-up detected. Closing it.")  # Debug: Log message when pop-up is detected
                pyautogui.press('enter')
                time.sleep(interval)
                return True
        except Exception as e:
            logging.error(f"Error checking pop-up: {e}")  # Debug
        time.sleep(interval)
    logging.debug("Pop-up not detected.")  # Debug: Log message when pop-up is not detected
    return False

# Function to wait until the mouse is no longer busy
def wait_until_not_busy():
    while True:
        cursor = win32gui.GetCursorInfo()[1]
        if cursor != win32api.LoadCursor(0, 32514):  # 32514 is the OCR_WAIT cursor
            break
        time.sleep(0.1)

# Global flag to stop automation
stop_automation_flag = False

# Function to stop the automation process
def stop_automation():
    global stop_automation_flag
    stop_automation_flag = True
    logging.debug("Stop automation flag set.")  # Debug: Log when stop is requested

# Function to update the listbox with the current data
def update_listbox(data):
    listbox.delete(0, tk.END)
    for index, row in data.iterrows():
        listbox.insert(tk.END, f"Date: {row['Date']}, Invoice: {row['Invoice Number']}, Serial: {row['Serial'] if 'Serial' in row and not pd.isna(row['Serial']) else ''}")

# Function to automate data entry
def automate_entry():
    global stop_automation_flag
    stop_automation_flag = False  # Ensure the flag is reset at the start
    try:
        logging.debug("Starting automation...")  # Debug
        if data is None:
            messagebox.showerror("No Data", "Please load the data first.")
            return

        # Ensure data is not empty
        if data.empty:
            logging.warning("Data is empty. Automation cannot proceed.")
            messagebox.showwarning("No Data", "The loaded data is empty. Please check the file content.")
            return

        def monitor_stop_key():
            keyboard.wait('`')
            stop_automation()

        monitor_thread = threading.Thread(target=monitor_stop_key)
        monitor_thread.start()

        for index, row in data.iterrows():
            if stop_automation_flag:
                logging.debug("Stop automation flag detected. Ending automation.")
                break

            logging.debug(f"Processing row {index}: {row}")  # Debug: Log the current row being processed
            user_date = row['Date']
            invoice_number = row['Invoice Number']
            serial_number = row['Serial'] if 'Serial' in row and not pd.isna(row['Serial']) else ''

            if pd.isna(invoice_number) or pd.isna(user_date):
                logging.debug(f"Skipping row {index} due to missing data")  # Debug: Log message if data is missing
                continue

            # Convert invoice_number to string if it's a number
            invoice_number = str(invoice_number)
            logging.debug(f"Invoice Number: {invoice_number}")  # Debug: Log the invoice number

            try:
                # Focus on the application window
                logging.debug("Focusing on application window...")  # Debug
                pyautogui.click(624, 229)  # Click to focus on the application window (coordinates may need adjustment)
                wait_until_not_busy()  # Wait for the mouse to stop being busy

                # Click on the "Order #" field (coordinates may need adjustment)
                logging.debug("Clicking on Order # field...")  # Debug
                pyautogui.click(689, 365)  # Adjust these coordinates to the "Order #" field
                wait_until_not_busy()
                pyautogui.write(invoice_number)  # Enter the invoice number
                logging.debug(f"Entered Invoice Number: {invoice_number}")  # Debug: Confirm invoice number entry
                pyautogui.press('tab')
                wait_until_not_busy()

                # Check for the pop-up and close it if it appears
                logging.debug("Checking for pop-up...")  # Debug
                check_and_close_popup('credit_limit_exceeded.png', retries=2, interval=1)
                wait_until_not_busy()

                # Change Mode/Prev/Disp: O to I, continue regardless of pop-up detection
                logging.debug("Changing Mode/Prev/Disp to I...")  # Debug
                pyautogui.click(660, 360)  # Ensure correct field is focused
                pyautogui.press('i')
                pyautogui.press('tab')
                time.sleep(1)
                pyautogui.press('enter')
                wait_until_not_busy()  # Wait for the new window to appear

                # If there is a serial number in column C
                if serial_number:
                    logging.debug(f"Entering Serial Number: {serial_number}")  # Debug
                    wait_until_not_busy()
                    pyautogui.write(serial_number)
                    wait_until_not_busy()
                    pyautogui.hotkey('alt', 'r')  # Close the screen with ALT+R
                    time.sleep(3)

                # Enter the date in the "Ship Date" field on the second screen
                logging.debug("Entering Ship Date...")  # Debug
                pyautogui.click(661, 522)  # Adjust these coordinates to the "Ship Date" field
                wait_until_not_busy()
                pyautogui.write(user_date)
                logging.debug(f"Entered Ship Date: {user_date}")  # Debug: Confirm date entry
                pyautogui.click(1169, 795)
                time.sleep(3)
                pyautogui.press('enter')
                time.sleep(1)

                # Log progress
                logging.debug("Automation step completed successfully.")  # Debug

                # Delete the processed row
                data.drop(index, inplace=True)
                data.reset_index(drop=True, inplace=True)
                data.to_excel(file_path, index=False, header=False)
                logging.debug("Deleted processed row and updated Excel file.")  # Debug

                # Update the listbox
                update_listbox(data)

            except Exception as automation_error:
                logging.error(f"Error during automation step for row {index}: {automation_error}")  # Debug
                break

    except Exception as e:
        logging.error(f"Error during automation: {e}")  # Debug: Log the exception
        messagebox.showerror("Error", f"Failed to complete automation: {e}")

    finally:
        monitor_thread.join()

def start_automation_thread():
    automation_thread = threading.Thread(target=automate_entry)
    automation_thread.start()

# Function to delete the selected line from the Excel sheet
def delete_selected_line():
    try:
        global data
        selected_index = listbox.curselection()
        if selected_index:
            index = selected_index[0]
            data.drop(index, inplace=True)  # Delete the selected row
            data.reset_index(drop=True, inplace=True)  # Reset the index
            data.to_excel(file_path, index=False, header=False)  # Save changes to the Excel file
            logging.debug("Deleted selected line from Excel file.")  # Debug
            update_listbox(data)  # Update the listbox
            messagebox.showinfo("Delete Complete", "Selected line has been deleted from the Excel sheet.")
        else:
            messagebox.showwarning("No Selection", "Please select a line to delete.")
    except Exception as e:
        logging.error(f"Error deleting selected line: {e}")  # Debug
        messagebox.showerror("Error", f"Failed to delete selected line: {e}")

# Function to purge data from the Excel sheet
def purge_data():
    try:
        global data
        data = pd.DataFrame(columns=['Date', 'Invoice Number', 'Serial'])
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Clear all rows, including headers
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
                    
            workbook.save(file_path)
            messagebox.showinfo("Purge Complete", "All data has been purged from the Excel sheet.")
            listbox.delete(0, tk.END)  # Clear the listbox
        else:
            messagebox.showinfo("File Not Found", "The Excel file does not exist.")
    except Exception as e:
        logging.error(f"Error purging data: {e}")  # Debug
        messagebox.showerror("Error", f"Failed to purge data: {e}")

# Function to write data to the Excel sheet
def write_to_excel(event=None):
    user_date = date_entry.get()
    if not user_date:
        messagebox.showerror("No Date", "Please enter the date first.")
        return

    try:
        # Format the date as MM/DD/YY
        formatted_date = datetime.strptime(user_date, "%Y-%m-%d").strftime("%m/%d/%y")
    except ValueError as e:
        logging.error(f"Error parsing date: {e}")  # Debug
        messagebox.showerror("Invalid Date", "Please enter the date in the format YYYY-MM-DD.")
        return

    invoice_number = invoice_entry.get()
    if not invoice_number:
        messagebox.showerror("No Invoice Number", "Please enter the invoice number.")
        return

    serial_value = serial_var.get()
    serial_text = ""

    if serial_value:
        serial_text = simpledialog.askstring("Serial Number", "Please enter the serial number:")
        if not serial_text:
            messagebox.showerror("No Serial Number", "Please enter the serial number.")
            return

    try:
        global data
        row_data = [formatted_date, invoice_number, serial_text] if serial_value else [formatted_date, invoice_number, '']

        if data is None or data.empty:
            data = pd.DataFrame([row_data], columns=['Date', 'Invoice Number', 'Serial'])
        else:
            data.loc[len(data)] = row_data

        data.to_excel(file_path, index=False, header=False)
        logging.debug(f"Writing row: {row_data}")  # Debug: Log the row data being written

        update_listbox(data)  # Update the listbox with the new data

        invoice_entry.delete(0, tk.END)
        serial_checkbox.deselect()
    except Exception as e:
        logging.error(f"Error writing to Excel: {e}")  # Debug
        messagebox.showerror("Error", f"Failed to write to Excel: {e}")

# Main GUI setup
root = tk.Tk()
root.title("Data Entry Automation")

data = None

# Create and place the date entry widget
date_label = tk.Label(root, text="Enter the date (YYYY-MM-DD):")
date_label.pack(pady=5)
date_entry = tk.Entry(root)
date_entry.pack(pady=5)

# Set default date to today's date
today_date = datetime.today().strftime("%Y-%m-%d")
date_entry.insert(0, today_date)

# Create and place the invoice number entry widget
invoice_label = tk.Label(root, text="Enter the invoice number:")
invoice_label.pack(pady=5)
invoice_entry = tk.Entry(root)
invoice_entry.pack(pady=5)

# Create and place the serial checkbox
serial_var = tk.IntVar()
serial_checkbox = tk.Checkbutton(root, text="Serial", variable=serial_var)
serial_checkbox.pack(pady=5)

# Create and place the load data button
load_button = tk.Button(root, text="Load Data", command=load_data)
load_button.pack(pady=10)

# Create and place the start automation button
start_button = tk.Button(root, text="Start Automation", command=start_automation_thread)
start_button.pack(pady=10)

# Create and place the delete selected line button
delete_selected_line_button = tk.Button(root, text="Delete", command=delete_selected_line)
delete_selected_line_button.pack(pady=10)

# Create and place the purge data button
purge_button = tk.Button(root, text="Purge Data", command=purge_data)
purge_button.pack(pady=10)

# Create and place the listbox to display progress
listbox = tk.Listbox(root, width=50, height=10)
listbox.pack(pady=10)

# Bind the Enter key to write to Excel
root.bind('<Return>', write_to_excel)

# Load data into the listbox when the program starts
load_data()

# Run the application
root.mainloop()
